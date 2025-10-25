import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelService:
    def __init__(self):
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.subheader_font = Font(bold=True, color="FFFFFF", size=11)
        self.left_seat_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        self.right_seat_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate_allocation_report(self, allocation):
        wb = Workbook()
        wb.remove(wb.active)

        self._create_summary_sheet(wb, allocation)

        allocations = allocation.get('allocations', [])
        for room_alloc in allocations:
            self._create_room_sheet(wb, room_alloc, allocation)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def generate_multi_exam_report(self, allocation):
        wb = Workbook()
        wb.remove(wb.active)

        self._create_multi_exam_summary_sheet(wb, allocation)

        allocations = allocation.get('allocations', [])
        for room_alloc in allocations:
            self._create_multi_exam_room_sheet(wb, room_alloc, allocation)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def generate_class_specific_report(self, allocation, class_year):
        wb = Workbook()
        wb.remove(wb.active)

        filtered_allocations = []
        allocations = allocation.get('allocations', [])

        for room_alloc in allocations:
            room = room_alloc['room']
            students = room_alloc['students']
            class_students = [
                student_alloc for student_alloc in students
                if str(student_alloc['student'].get('year', '')) == str(class_year)
            ]

            if class_students:
                filtered_allocations.append({
                    'room': room,
                    'students': class_students,
                    'subject_breakdown': room_alloc.get('subject_breakdown', {}),
                    'total_in_room': len(students)
                })

        self._create_class_summary_sheet(wb, allocation, class_year, filtered_allocations)

        for room_alloc in filtered_allocations:
            self._create_class_room_sheet(wb, room_alloc, class_year)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _create_summary_sheet(self, wb, allocation):
        ws = wb.create_sheet("Summary")

        ws['A1'] = "Exam Seat Allocation Report"
        ws['A1'].font = Font(bold=True, size=16, color="1F4E78")
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')

        row = 3
        ws[f'A{row}'] = "Generated On:"
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Strategy:"
        ws[f'B{row}'] = allocation.get('strategy', 'N/A').title()
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Subject Filter:"
        ws[f'B{row}'] = allocation.get('subject_filter', 'All Subjects') or 'All Subjects'
        ws[f'A{row}'].font = Font(bold=True)

        summary = allocation.get('allocation_summary', {})
        row += 2
        ws[f'A{row}'] = "Allocation Summary"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="1F4E78")
        ws.merge_cells(f'A{row}:B{row}')

        row += 1
        summary_data = [
            ('Total Students:', summary.get('total_students', 0)),
            ('Students Allocated:', summary.get('total_allocated', 0)),
            ('Students Unallocated:', summary.get('total_unallocated', 0)),
            ('Rooms Used:', summary.get('rooms_used', 0)),
            ('Allocation Rate:', f"{summary.get('allocation_percentage', 0)}%"),
            ('Quality Rating:', summary.get('quality_rating', 'N/A')),
        ]

        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30

    def _create_room_sheet(self, wb, room_alloc, allocation):
        room = room_alloc['room']
        students = room_alloc['students']

        sheet_name = room['name'][:31].replace('/', '-').replace('\\', '-').replace('*', '')
        ws = wb.create_sheet(sheet_name)

        ws['A1'] = f"Room: {room['name']}"
        ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')

        room_layout = room_alloc.get('room_layout', {})
        ws['A2'] = f"Capacity: {room['capacity']} | Allocated: {len(students)} | Benches: {(len(students) + 1) // 2}"
        ws.merge_cells('A2:F2')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A2'].font = Font(italic=True)

        if room_layout:
            ws['A3'] = f"Layout: {room_layout.get('total_rows', 0)} rows × {room_layout.get('total_columns', 0)} columns | {room_layout.get('benches_per_row', 0)} benches per row"
            ws.merge_cells('A3:F3')
            ws['A3'].alignment = Alignment(horizontal='center')
            ws['A3'].font = Font(italic=True, size=9, color="666666")

        subject_breakdown = room_alloc.get('subject_breakdown', {})
        if subject_breakdown:
            ws['A4'] = "Subject Distribution: " + " | ".join([f"{subj}: {count}" for subj, count in subject_breakdown.items()])
            ws.merge_cells('A4:F4')
            ws['A4'].alignment = Alignment(horizontal='center')
            ws['A4'].font = Font(italic=True, size=9)

        row = 6
        headers = ['Bench #', 'Left Seat #', 'Left Roll #', 'Right Seat #', 'Right Roll #']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num, value=header)
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border

        sorted_students = sorted(students, key=lambda x: x['seat_number'])

        row += 1
        bench_num = 1
        for i in range(0, len(sorted_students), 2):
            left_student = sorted_students[i]
            right_student = sorted_students[i + 1] if i + 1 < len(sorted_students) else None

            ws.cell(row=row, column=1, value=bench_num).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=1).border = self.border
            ws.cell(row=row, column=1).font = Font(bold=True)

            ws.cell(row=row, column=2, value=left_student['seat_number']).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=2).fill = self.left_seat_fill
            ws.cell(row=row, column=2).border = self.border
            ws.cell(row=row, column=2).font = Font(bold=True)

            ws.cell(row=row, column=3, value=left_student['student']['roll_number'])
            ws.cell(row=row, column=3).fill = self.left_seat_fill
            ws.cell(row=row, column=3).border = self.border

            ws.cell(row=row, column=4, value=left_student['student']['roll_number']).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=4).fill = self.left_seat_fill
            ws.cell(row=row, column=4).border = self.border

            if right_student:
                ws.cell(row=row, column=5, value=right_student['seat_number']).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row, column=5).fill = self.right_seat_fill
                ws.cell(row=row, column=5).border = self.border
                ws.cell(row=row, column=5).font = Font(bold=True)

                ws.cell(row=row, column=6, value=right_student['student']['roll_number'])
                ws.cell(row=row, column=6).fill = self.right_seat_fill
                ws.cell(row=row, column=6).border = self.border

                ws.cell(row=row, column=7, value=right_student['student']['roll_number']).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=7).fill = self.right_seat_fill
                ws.cell(row=row, column=7).border = self.border
            else:
                for col in range(5, 8):
                    ws.cell(row=row, column=col, value='')
                    ws.cell(row=row, column=col).fill = self.right_seat_fill
                    ws.cell(row=row, column=col).border = self.border

            row += 1
            bench_num += 1

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15

        if students and students[0].get('grid'):
            self._create_grid_visualization_sheet(wb, room_alloc)

    def _create_grid_visualization_sheet(self, wb, room_alloc):
        room = room_alloc['room']
        students = room_alloc['students']
        room_layout = room_alloc.get('room_layout', {})

        sheet_name = f"{room['name'][:25]}_Grid".replace('/', '-').replace('\\', '-').replace('*', '')
        ws = wb.create_sheet(sheet_name)

        ws['A1'] = f"Grid Layout - {room['name']}"
        ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')

        total_rows = room_layout.get('total_rows', 0)
        total_cols = room_layout.get('total_columns', 0)
        benches_per_row = room_layout.get('benches_per_row', 4)

        ws['A2'] = f"Layout: {total_rows} rows × {total_cols} columns | {benches_per_row} benches per row | 2 seats per bench"
        ws.merge_cells('A2:F2')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A2'].font = Font(italic=True)

        grid_map = {}
        for student_alloc in students:
            grid = student_alloc.get('grid', {})
            row = grid.get('row', 0)
            col = grid.get('col', 0)
            position = grid.get('position', 'left')

            if row not in grid_map:
                grid_map[row] = {}
            if col not in grid_map[row]:
                grid_map[row][col] = {'left': None, 'right': None}

            grid_map[row][col][position] = student_alloc

        start_row = 4
        current_row = start_row

        ws.cell(row=current_row, column=1, value="Row\\Col").font = Font(bold=True)
        ws.cell(row=current_row, column=1).fill = self.header_fill
        ws.cell(row=current_row, column=1).font = self.header_font
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')

        for col_idx in range(total_cols):
            excel_col = 2 + (col_idx * 2)
            ws.merge_cells(start_row=current_row, start_column=excel_col, end_row=current_row, end_column=excel_col + 1)
            cell = ws.cell(row=current_row, column=excel_col, value=f"Col {col_idx}")
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        current_row += 1

        ws.cell(row=current_row, column=1, value="").fill = self.subheader_fill
        for col_idx in range(total_cols):
            excel_col = 2 + (col_idx * 2)

            cell_l = ws.cell(row=current_row, column=excel_col, value="L")
            cell_l.font = self.subheader_font
            cell_l.fill = self.subheader_fill
            cell_l.alignment = Alignment(horizontal='center', vertical='center')
            cell_l.border = self.border

            cell_r = ws.cell(row=current_row, column=excel_col + 1, value="R")
            cell_r.font = self.subheader_font
            cell_r.fill = self.subheader_fill
            cell_r.alignment = Alignment(horizontal='center', vertical='center')
            cell_r.border = self.border

        current_row += 1

        for row_idx in range(total_rows):
            cell = ws.cell(row=current_row, column=1, value=f"Row {row_idx}")
            cell.font = Font(bold=True)
            cell.fill = self.subheader_fill
            cell.font = self.subheader_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

            for col_idx in range(total_cols):
                excel_col = 2 + (col_idx * 2)

                bench = grid_map.get(row_idx, {}).get(col_idx, {'left': None, 'right': None})

                if bench['left']:
                    student = bench['left']['student']
                    subjects = student.get('subjects', [student.get('subject', 'N/A')])
                    primary_subject = subjects[0] if subjects else 'N/A'
                    cell_value = f"{student['roll_number']}\n{primary_subject[:8]}"

                    cell = ws.cell(row=current_row, column=excel_col, value=cell_value)
                    cell.fill = self.left_seat_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = self.border
                    cell.font = Font(size=9)
                else:
                    cell = ws.cell(row=current_row, column=excel_col, value="")
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    cell.border = self.border

                if bench['right']:
                    student = bench['right']['student']
                    subjects = student.get('subjects', [student.get('subject', 'N/A')])
                    primary_subject = subjects[0] if subjects else 'N/A'
                    cell_value = f"{student['roll_number']}\n{primary_subject[:8]}"

                    cell = ws.cell(row=current_row, column=excel_col + 1, value=cell_value)
                    cell.fill = self.right_seat_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = self.border
                    cell.font = Font(size=9)
                else:
                    cell = ws.cell(row=current_row, column=excel_col + 1, value="")
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    cell.border = self.border

            ws.row_dimensions[current_row].height = 35
            current_row += 1

        ws.column_dimensions['A'].width = 10
        for col_idx in range(total_cols * 2):
            col_letter = get_column_letter(2 + col_idx)
            ws.column_dimensions[col_letter].width = 12

    def _create_multi_exam_summary_sheet(self, wb, allocation):
        ws = wb.create_sheet("Summary")

        ws['A1'] = "Multi-Exam Seat Allocation Report"
        ws['A1'].font = Font(bold=True, size=16, color="1F4E78")
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')

        session_info = allocation.get('session_info', {})
        summary = allocation.get('allocation_summary', {})

        row = 3
        ws[f'A{row}'] = "Session Name:"
        ws[f'B{row}'] = session_info.get('session_name', 'Multi-Exam Session')
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Generated On:"
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f'A{row}'].font = Font(bold=True)

        row += 2
        ws[f'A{row}'] = "Allocation Summary"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="1F4E78")

        row += 1
        summary_data = [
            ('Total Students:', summary.get('total_students', 0)),
            ('Students Allocated:', summary.get('total_allocated', 0)),
            ('Rooms Used:', summary.get('rooms_used', 0)),
            ('Allocation Rate:', f"{summary.get('allocation_percentage', 0)}%"),
        ]

        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30

    def _create_multi_exam_room_sheet(self, wb, room_alloc, allocation):
        room = room_alloc['room']
        students = room_alloc['students']

        sheet_name = room['name'][:31].replace('/', '-').replace('\\', '-').replace('*', '')
        ws = wb.create_sheet(sheet_name)

        ws['A1'] = f"Room: {room['name']}"
        ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A2'] = f"Capacity: {room['capacity']} | Allocated: {len(students)}"
        ws.merge_cells('A2:H2')
        ws['A2'].alignment = Alignment(horizontal='center')

        row = 4
        headers = ['Bench #', 'Left Seat #', 'Left Roll #', 'Left Exam', 'Right Seat #', 'Right Roll #', 'Right Exam']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num, value=header)
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border

        sorted_students = sorted(students, key=lambda x: x['seat_number'])

        row += 1
        bench_num = 1
        for i in range(0, len(sorted_students), 2):
            left_student = sorted_students[i]
            right_student = sorted_students[i + 1] if i + 1 < len(sorted_students) else None

            ws.cell(row=row, column=1, value=bench_num).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=1).border = self.border
            ws.cell(row=row, column=1).font = Font(bold=True)

            ws.cell(row=row, column=2, value=left_student['seat_number']).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=2).fill = self.left_seat_fill
            ws.cell(row=row, column=2).border = self.border
            ws.cell(row=row, column=2).font = Font(bold=True)

            ws.cell(row=row, column=3, value=left_student['student']['roll_number']).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3).fill = self.left_seat_fill
            ws.cell(row=row, column=3).border = self.border

            ws.cell(row=row, column=4, value=left_student['student'].get('exam_name', 'N/A'))
            ws.cell(row=row, column=4).fill = self.left_seat_fill
            ws.cell(row=row, column=4).border = self.border

            if right_student:
                ws.cell(row=row, column=5, value=right_student['seat_number']).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=5).fill = self.right_seat_fill
                ws.cell(row=row, column=5).border = self.border
                ws.cell(row=row, column=5).font = Font(bold=True)

                ws.cell(row=row, column=6, value=right_student['student']['roll_number']).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=6).fill = self.right_seat_fill
                ws.cell(row=row, column=6).border = self.border

                ws.cell(row=row, column=7, value=right_student['student'].get('exam_name', 'N/A'))
                ws.cell(row=row, column=7).fill = self.right_seat_fill
                ws.cell(row=row, column=7).border = self.border
            else:
                for col in range(5, 8):
                    ws.cell(row=row, column=col, value='')
                    ws.cell(row=row, column=col).fill = self.right_seat_fill
                    ws.cell(row=row, column=col).border = self.border

            row += 1
            bench_num += 1

        for i, width in enumerate([10, 12, 15, 15, 12, 15, 15], 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _create_class_summary_sheet(self, wb, allocation, class_year, filtered_allocations):
        ws = wb.create_sheet("Summary")

        ws['A1'] = f"Class {class_year} Seat Allocation Report"
        ws['A1'].font = Font(bold=True, size=16, color="1F4E78")
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')

        row = 3
        ws[f'A{row}'] = "Generated On:"
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Class/Year:"
        ws[f'B{row}'] = f"{class_year}{'st' if class_year == '1' else 'nd' if class_year == '2' else 'rd' if class_year == '3' else 'th'} Year"
        ws[f'A{row}'].font = Font(bold=True)

        row += 2
        total_students = sum(len(alloc['students']) for alloc in filtered_allocations)
        ws[f'A{row}'] = "Class Summary"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="1F4E78")

        row += 1
        summary_data = [
            ('Total Class Students:', total_students),
            ('Rooms with Class Students:', len(filtered_allocations)),
        ]

        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30

    def _create_class_room_sheet(self, wb, room_alloc, class_year):
        room = room_alloc['room']
        students = room_alloc['students']

        sheet_name = f"{room['name'][:25]}-C{class_year}".replace('/', '-').replace('\\', '-').replace('*', '')
        ws = wb.create_sheet(sheet_name)

        ws['A1'] = f"Room: {room['name']} - Class {class_year}"
        ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A2'] = f"Class {class_year} Students: {len(students)}/{room_alloc.get('total_in_room', len(students))}"
        ws.merge_cells('A2:G2')
        ws['A2'].alignment = Alignment(horizontal='center')

        row = 4
        headers = ['Bench #', 'Left Seat #', 'Left Roll #', 'Right Seat #', 'Right Roll #']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num, value=header)
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border

        sorted_students = sorted(students, key=lambda x: x['seat_number'])

        row += 1
        bench_num = 1
        for i in range(0, len(sorted_students), 2):
            left_student = sorted_students[i]
            right_student = sorted_students[i + 1] if i + 1 < len(sorted_students) else None

            ws.cell(row=row, column=1, value=bench_num).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=1).border = self.border
            ws.cell(row=row, column=1).font = Font(bold=True)

            ws.cell(row=row, column=2, value=left_student['seat_number']).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2).fill = self.left_seat_fill
            ws.cell(row=row, column=2).border = self.border
            ws.cell(row=row, column=2).font = Font(bold=True)

            ws.cell(row=row, column=3, value=left_student['student']['roll_number']).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3).fill = self.left_seat_fill
            ws.cell(row=row, column=3).border = self.border

            if right_student:
                ws.cell(row=row, column=4, value=right_student['seat_number']).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=4).fill = self.right_seat_fill
                ws.cell(row=row, column=4).border = self.border
                ws.cell(row=row, column=4).font = Font(bold=True)

                ws.cell(row=row, column=5, value=right_student['student']['roll_number']).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=5).fill = self.right_seat_fill
                ws.cell(row=row, column=5).border = self.border
            else:
                for col in range(4, 6):
                    ws.cell(row=row, column=col, value='')
                    ws.cell(row=row, column=col).fill = self.right_seat_fill
                    ws.cell(row=row, column=col).border = self.border

            row += 1
            bench_num += 1

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
